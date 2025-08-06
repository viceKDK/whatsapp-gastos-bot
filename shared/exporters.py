"""
Sistema de Exportación Multi-formato

Exporta datos de gastos a múltiples formatos: Excel, CSV, PDF, JSON, HTML.
"""

import json
import csv
from pathlib import Path
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Union
from decimal import Decimal
from dataclasses import dataclass
from enum import Enum
import tempfile

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

from domain.models.gasto import Gasto
from config.config_manager import get_config
from shared.logger import get_logger


logger = get_logger(__name__)


class ExportFormat(Enum):
    """Formatos de exportación soportados."""
    EXCEL = "excel"
    CSV = "csv"
    PDF = "pdf"
    JSON = "json"
    HTML = "html"


@dataclass
class ExportOptions:
    """Opciones de exportación."""
    include_charts: bool = True
    include_summary: bool = True
    date_format: str = "%Y-%m-%d"
    group_by_category: bool = True
    group_by_month: bool = False
    filter_start_date: Optional[date] = None
    filter_end_date: Optional[date] = None
    custom_title: Optional[str] = None
    
    def to_dict(self) -> Dict[str, Any]:
        """Convierte a diccionario."""
        return {
            'include_charts': self.include_charts,
            'include_summary': self.include_summary,
            'date_format': self.date_format,
            'group_by_category': self.group_by_category,
            'group_by_month': self.group_by_month,
            'filter_start_date': self.filter_start_date.isoformat() if self.filter_start_date else None,
            'filter_end_date': self.filter_end_date.isoformat() if self.filter_end_date else None,
            'custom_title': self.custom_title
        }


@dataclass
class ExportResult:
    """Resultado de exportación."""
    success: bool
    file_path: Optional[str] = None
    file_size_bytes: int = 0
    record_count: int = 0
    format: Optional[ExportFormat] = None
    error_message: Optional[str] = None
    export_time_seconds: float = 0
    metadata: Dict[str, Any] = None
    
    def __post_init__(self):
        if self.metadata is None:
            self.metadata = {}


class BaseExporter:
    """Clase base para exportadores."""
    
    def __init__(self):
        self.config = get_config()
        self.logger = logger
    
    def export(self, gastos: List[Gasto], output_path: str, 
               options: ExportOptions = None) -> ExportResult:
        """
        Exporta gastos al formato específico.
        
        Args:
            gastos: Lista de gastos a exportar
            output_path: Ruta del archivo de salida
            options: Opciones de exportación
            
        Returns:
            Resultado de la exportación
        """
        raise NotImplementedError
    
    def _filter_gastos(self, gastos: List[Gasto], 
                      options: ExportOptions) -> List[Gasto]:
        """Filtra gastos según opciones."""
        if not options:
            return gastos
        
        filtered = gastos
        
        # Filtrar por fecha
        if options.filter_start_date or options.filter_end_date:
            start_date = options.filter_start_date or date.min
            end_date = options.filter_end_date or date.max
            
            filtered = [
                g for g in filtered 
                if start_date <= g.fecha.date() <= end_date
            ]
        
        return filtered
    
    def _calculate_summary(self, gastos: List[Gasto]) -> Dict[str, Any]:
        """Calcula resumen estadístico de gastos."""
        if not gastos:
            return {
                'total_gastos': 0,
                'monto_total': 0,
                'monto_promedio': 0,
                'categorias': {},
                'fecha_primer_gasto': None,
                'fecha_ultimo_gasto': None
            }
        
        from collections import defaultdict
        import statistics
        
        # Estadísticas básicas
        montos = [float(g.monto) for g in gastos]
        total_gastos = len(gastos)
        monto_total = sum(montos)
        monto_promedio = statistics.mean(montos)
        
        # Agrupar por categoría
        categorias = defaultdict(list)
        for gasto in gastos:
            categorias[gasto.categoria].append(float(gasto.monto))
        
        categoria_stats = {}
        for categoria, montos_cat in categorias.items():
            categoria_stats[categoria] = {
                'cantidad': len(montos_cat),
                'total': sum(montos_cat),
                'promedio': statistics.mean(montos_cat),
                'porcentaje': (sum(montos_cat) / monto_total * 100) if monto_total > 0 else 0
            }
        
        # Fechas extremas
        fechas = [g.fecha for g in gastos]
        fecha_primer_gasto = min(fechas)
        fecha_ultimo_gasto = max(fechas)
        
        return {
            'total_gastos': total_gastos,
            'monto_total': monto_total,
            'monto_promedio': monto_promedio,
            'monto_maximo': max(montos),
            'monto_minimo': min(montos),
            'categorias': categoria_stats,
            'fecha_primer_gasto': fecha_primer_gasto,
            'fecha_ultimo_gasto': fecha_ultimo_gasto
        }


class ExcelExporter(BaseExporter):
    """Exportador a formato Excel."""
    
    def export(self, gastos: List[Gasto], output_path: str, 
               options: ExportOptions = None) -> ExportResult:
        """Exporta a Excel con formato avanzado."""
        if not HAS_OPENPYXL:
            return ExportResult(
                success=False,
                error_message="openpyxl no está instalado"
            )
        
        start_time = datetime.now()
        
        try:
            options = options or ExportOptions()
            filtered_gastos = self._filter_gastos(gastos, options)
            
            if not filtered_gastos:
                return ExportResult(
                    success=False,
                    error_message="No hay gastos para exportar"
                )
            
            # Crear workbook
            wb = Workbook()
            
            # Hoja principal de gastos
            ws_gastos = wb.active
            ws_gastos.title = "Gastos"
            
            self._create_gastos_sheet(ws_gastos, filtered_gastos, options)
            
            # Hoja de resumen si está habilitado
            if options.include_summary:
                ws_resumen = wb.create_sheet("Resumen")
                summary = self._calculate_summary(filtered_gastos)
                self._create_summary_sheet(ws_resumen, summary, options)
            
            # Hoja de gráficos si está habilitado
            if options.include_charts:
                ws_graficos = wb.create_sheet("Gráficos")
                self._create_charts_sheet(ws_graficos, filtered_gastos, options)
            
            # Guardar archivo
            wb.save(output_path)
            
            # Calcular resultado
            file_size = Path(output_path).stat().st_size
            export_time = (datetime.now() - start_time).total_seconds()
            
            return ExportResult(
                success=True,
                file_path=output_path,
                file_size_bytes=file_size,
                record_count=len(filtered_gastos),
                format=ExportFormat.EXCEL,
                export_time_seconds=export_time
            )
            
        except Exception as e:
            self.logger.error(f"Error exportando a Excel: {e}")
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time_seconds=(datetime.now() - start_time).total_seconds()
            )
    
    def _create_gastos_sheet(self, ws, gastos: List[Gasto], options: ExportOptions):
        """Crea hoja de gastos detallados."""
        # Headers
        headers = ['Fecha', 'Hora', 'Monto', 'Categoría', 'Descripción']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Datos
        for row, gasto in enumerate(gastos, 2):
            ws.cell(row=row, column=1, value=gasto.fecha.strftime(options.date_format))
            ws.cell(row=row, column=2, value=gasto.fecha.strftime("%H:%M:%S"))
            ws.cell(row=row, column=3, value=float(gasto.monto))
            ws.cell(row=row, column=4, value=gasto.categoria)
            ws.cell(row=row, column=5, value=gasto.descripcion or "")
            
            # Formato de monto
            ws.cell(row=row, column=3).number_format = '"$"#,##0.00'
        
        # Ajustar ancho de columnas
        column_widths = [12, 10, 12, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Agregar total al final
        total_row = len(gastos) + 2
        ws.cell(row=total_row, column=2, value="TOTAL:").font = Font(bold=True)
        total_formula = f"=SUM(C2:C{len(gastos) + 1})"
        total_cell = ws.cell(row=total_row, column=3, value=total_formula)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '"$"#,##0.00'
        total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    def _create_summary_sheet(self, ws, summary: Dict[str, Any], options: ExportOptions):
        """Crea hoja de resumen estadístico."""
        # Título
        ws.cell(row=1, column=1, value="RESUMEN DE GASTOS").font = Font(bold=True, size=16)
        
        row = 3
        
        # Estadísticas generales
        ws.cell(row=row, column=1, value="Estadísticas Generales").font = Font(bold=True)
        row += 1
        
        stats = [
            ("Total de gastos:", summary['total_gastos']),
            ("Monto total:", f"${summary['monto_total']:,.2f}"),
            ("Monto promedio:", f"${summary['monto_promedio']:,.2f}"),
            ("Monto máximo:", f"${summary['monto_maximo']:,.2f}"),
            ("Monto mínimo:", f"${summary['monto_minimo']:,.2f}"),
            ("Período:", f"{summary['fecha_primer_gasto'].strftime(options.date_format)} - {summary['fecha_ultimo_gasto'].strftime(options.date_format)}")
        ]
        
        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 2
        
        # Resumen por categoría
        ws.cell(row=row, column=1, value="Resumen por Categoría").font = Font(bold=True)
        row += 1
        
        # Headers de tabla de categorías
        cat_headers = ['Categoría', 'Cantidad', 'Total', 'Promedio', '%']
        for col, header in enumerate(cat_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        row += 1
        
        # Datos de categorías
        for categoria, stats in summary['categorias'].items():
            ws.cell(row=row, column=1, value=categoria)
            ws.cell(row=row, column=2, value=stats['cantidad'])
            ws.cell(row=row, column=3, value=f"${stats['total']:,.2f}")
            ws.cell(row=row, column=4, value=f"${stats['promedio']:,.2f}")
            ws.cell(row=row, column=5, value=f"{stats['porcentaje']:.1f}%")
            row += 1
    
    def _create_charts_sheet(self, ws, gastos: List[Gasto], options: ExportOptions):
        """Crea hoja con gráficos."""
        # Preparar datos para gráficos
        from collections import defaultdict
        
        categorias = defaultdict(float)
        for gasto in gastos:
            categorias[gasto.categoria] += float(gasto.monto)
        
        # Crear datos para gráfico circular
        ws.cell(row=1, column=1, value="Categoría")
        ws.cell(row=1, column=2, value="Monto")
        
        for row, (categoria, monto) in enumerate(categorias.items(), 2):
            ws.cell(row=row, column=1, value=categoria)
            ws.cell(row=row, column=2, value=monto)
        
        # Crear gráfico circular
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(categorias) + 1)
        data = Reference(ws, min_col=2, min_row=2, max_row=len(categorias) + 1)
        pie.add_data(data)
        pie.set_categories(labels)
        pie.title = "Gastos por Categoría"
        
        ws.add_chart(pie, "D2")


class CSVExporter(BaseExporter):
    """Exportador a formato CSV."""
    
    def export(self, gastos: List[Gasto], output_path: str, 
               options: ExportOptions = None) -> ExportResult:
        """Exporta a CSV."""
        start_time = datetime.now()
        
        try:
            options = options or ExportOptions()
            filtered_gastos = self._filter_gastos(gastos, options)
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['fecha', 'hora', 'monto', 'categoria', 'descripcion']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                
                for gasto in filtered_gastos:
                    writer.writerow({
                        'fecha': gasto.fecha.strftime(options.date_format),
                        'hora': gasto.fecha.strftime("%H:%M:%S"),
                        'monto': float(gasto.monto),
                        'categoria': gasto.categoria,
                        'descripcion': gasto.descripcion or ""
                    })
            
            file_size = Path(output_path).stat().st_size
            export_time = (datetime.now() - start_time).total_seconds()
            
            return ExportResult(
                success=True,
                file_path=output_path,
                file_size_bytes=file_size,
                record_count=len(filtered_gastos),
                format=ExportFormat.CSV,
                export_time_seconds=export_time
            )
            
        except Exception as e:
            self.logger.error(f"Error exportando a CSV: {e}")
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time_seconds=(datetime.now() - start_time).total_seconds()
            )


class JSONExporter(BaseExporter):
    """Exportador a formato JSON."""
    
    def export(self, gastos: List[Gasto], output_path: str, 
               options: ExportOptions = None) -> ExportResult:
        """Exporta a JSON."""
        start_time = datetime.now()
        
        try:
            options = options or ExportOptions()
            filtered_gastos = self._filter_gastos(gastos, options)
            
            # Preparar datos
            export_data = {
                'metadata': {
                    'export_date': datetime.now().isoformat(),
                    'record_count': len(filtered_gastos),
                    'format_version': '1.0',
                    'options': options.to_dict()
                },
                'gastos': []
            }
            
            # Incluir resumen si está habilitado
            if options.include_summary:
                export_data['summary'] = self._calculate_summary(filtered_gastos)
            
            # Convertir gastos a formato JSON
            for gasto in filtered_gastos:
                gasto_dict = {
                    'id': gasto.id,
                    'monto': float(gasto.monto),
                    'categoria': gasto.categoria,
                    'descripcion': gasto.descripcion,
                    'fecha': gasto.fecha.isoformat(),
                    'fecha_formatted': gasto.fecha.strftime(options.date_format),
                    'hora': gasto.fecha.strftime("%H:%M:%S")
                }
                export_data['gastos'].append(gasto_dict)
            
            # Escribir archivo JSON
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            file_size = Path(output_path).stat().st_size
            export_time = (datetime.now() - start_time).total_seconds()
            
            return ExportResult(
                success=True,
                file_path=output_path,
                file_size_bytes=file_size,
                record_count=len(filtered_gastos),
                format=ExportFormat.JSON,
                export_time_seconds=export_time
            )
            
        except Exception as e:
            self.logger.error(f"Error exportando a JSON: {e}")
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time_seconds=(datetime.now() - start_time).total_seconds()
            )


class PDFExporter(BaseExporter):
    """Exportador a formato PDF."""
    
    def export(self, gastos: List[Gasto], output_path: str, 
               options: ExportOptions = None) -> ExportResult:
        """Exporta a PDF."""
        if not HAS_REPORTLAB:
            return ExportResult(
                success=False,
                error_message="reportlab no está instalado"
            )
        
        start_time = datetime.now()
        
        try:
            options = options or ExportOptions()
            filtered_gastos = self._filter_gastos(gastos, options)
            
            # Crear documento PDF
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Título
            title = options.custom_title or "Reporte de Gastos"
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Centrado
            )
            story.append(Paragraph(title, title_style))
            
            # Información del reporte
            info_text = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>"
            info_text += f"Total de gastos: {len(filtered_gastos)}<br/>"
            if options.filter_start_date:
                info_text += f"Desde: {options.filter_start_date.strftime(options.date_format)}<br/>"
            if options.filter_end_date:
                info_text += f"Hasta: {options.filter_end_date.strftime(options.date_format)}"
            
            story.append(Paragraph(info_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Resumen si está habilitado
            if options.include_summary:
                summary = self._calculate_summary(filtered_gastos)
                self._add_summary_to_pdf(story, summary, styles, options)
            
            # Tabla de gastos
            self._add_gastos_table_to_pdf(story, filtered_gastos, styles, options)
            
            # Gráficos si están habilitados
            if options.include_charts and HAS_MATPLOTLIB:
                self._add_charts_to_pdf(story, filtered_gastos, styles, options)
            
            # Construir PDF
            doc.build(story)
            
            file_size = Path(output_path).stat().st_size
            export_time = (datetime.now() - start_time).total_seconds()
            
            return ExportResult(
                success=True,
                file_path=output_path,
                file_size_bytes=file_size,
                record_count=len(filtered_gastos),
                format=ExportFormat.PDF,
                export_time_seconds=export_time
            )
            
        except Exception as e:
            self.logger.error(f"Error exportando a PDF: {e}")
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time_seconds=(datetime.now() - start_time).total_seconds()
            )
    
    def _add_summary_to_pdf(self, story, summary: Dict[str, Any], 
                           styles, options: ExportOptions):
        """Agrega resumen al PDF."""
        story.append(Paragraph("Resumen", styles['Heading2']))
        
        # Tabla de resumen
        summary_data = [
            ['Concepto', 'Valor'],
            ['Total de gastos', str(summary['total_gastos'])],
            ['Monto total', f"${summary['monto_total']:,.2f}"],
            ['Monto promedio', f"${summary['monto_promedio']:,.2f}"],
            ['Monto máximo', f"${summary['monto_maximo']:,.2f}"],
            ['Monto mínimo', f"${summary['monto_minimo']:,.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
    
    def _add_gastos_table_to_pdf(self, story, gastos: List[Gasto], 
                                styles, options: ExportOptions):
        """Agrega tabla de gastos al PDF."""
        story.append(Paragraph("Detalle de Gastos", styles['Heading2']))
        
        # Preparar datos de tabla
        table_data = [['Fecha', 'Monto', 'Categoría', 'Descripción']]
        
        for gasto in gastos:
            table_data.append([
                gasto.fecha.strftime(options.date_format),
                f"${float(gasto.monto):,.2f}",
                gasto.categoria,
                (gasto.descripcion or "")[:30] + "..." if (gasto.descripcion or "") and len(gasto.descripcion) > 30 else (gasto.descripcion or "")
            ])
        
        # Crear tabla
        table = Table(table_data, colWidths=[1.5*inch, 1.2*inch, 1.5*inch, 2.3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),  # Alinear montos a la derecha
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(table)
    
    def _add_charts_to_pdf(self, story, gastos: List[Gasto], 
                          styles, options: ExportOptions):
        """Agrega gráficos al PDF."""
        try:
            story.append(Spacer(1, 20))
            story.append(Paragraph("Gráficos", styles['Heading2']))
            
            # Preparar datos por categoría
            from collections import defaultdict
            categorias = defaultdict(float)
            for gasto in gastos:
                categorias[gasto.categoria] += float(gasto.monto)
            
            # Crear gráfico circular con matplotlib
            fig, ax = plt.subplots(figsize=(8, 6))
            labels = list(categorias.keys())
            values = list(categorias.values())
            
            ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
            ax.set_title('Distribución de Gastos por Categoría')
            
            # Guardar gráfico temporalmente
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                fig.savefig(tmp.name, format='png', bbox_inches='tight', dpi=150)
                temp_chart_path = tmp.name
            
            plt.close(fig)
            
            # Agregar imagen al PDF
            story.append(Image(temp_chart_path, width=6*inch, height=4*inch))
            
            # Limpiar archivo temporal
            Path(temp_chart_path).unlink()
            
        except Exception as e:
            self.logger.error(f"Error agregando gráficos al PDF: {e}")


class HTMLExporter(BaseExporter):
    """Exportador a formato HTML."""
    
    def export(self, gastos: List[Gasto], output_path: str, 
               options: ExportOptions = None) -> ExportResult:
        """Exporta a HTML."""
        start_time = datetime.now()
        
        try:
            options = options or ExportOptions()
            filtered_gastos = self._filter_gastos(gastos, options)
            
            html_content = self._generate_html(filtered_gastos, options)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            file_size = Path(output_path).stat().st_size
            export_time = (datetime.now() - start_time).total_seconds()
            
            return ExportResult(
                success=True,
                file_path=output_path,
                file_size_bytes=file_size,
                record_count=len(filtered_gastos),
                format=ExportFormat.HTML,
                export_time_seconds=export_time
            )
            
        except Exception as e:
            self.logger.error(f"Error exportando a HTML: {e}")
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time_seconds=(datetime.now() - start_time).total_seconds()
            )
    
    def _generate_html(self, gastos: List[Gasto], options: ExportOptions) -> str:
        """Genera contenido HTML."""
        title = options.custom_title or "Reporte de Gastos"
        
        html = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            text-align: center;
            margin-bottom: 30px;
        }}
        .info {{
            background-color: #ecf0f1;
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 30px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }}
        th {{
            background-color: #34495e;
            color: white;
            font-weight: bold;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .amount {{
            text-align: right;
            font-weight: bold;
        }}
        .total-row {{
            background-color: #3498db !important;
            color: white;
            font-weight: bold;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 30px;
        }}
        .summary-card {{
            background-color: #ecf0f1;
            padding: 15px;
            border-radius: 5px;
            text-align: center;
        }}
        .summary-card h3 {{
            margin: 0 0 10px 0;
            color: #2c3e50;
        }}
        .summary-card .value {{
            font-size: 1.5em;
            font-weight: bold;
            color: #27ae60;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        
        <div class="info">
            <strong>Generado el:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br>
            <strong>Total de gastos:</strong> {len(gastos)}<br>
            <strong>Período:</strong> {self._get_period_text(gastos, options)}
        </div>
"""
        
        # Agregar resumen si está habilitado
        if options.include_summary:
            summary = self._calculate_summary(gastos)
            html += self._generate_summary_html(summary)
        
        # Tabla de gastos
        html += self._generate_gastos_table_html(gastos, options)
        
        html += """
    </div>
</body>
</html>
"""
        
        return html
    
    def _generate_summary_html(self, summary: Dict[str, Any]) -> str:
        """Genera HTML del resumen."""
        return f"""
        <h2>Resumen</h2>
        <div class="summary">
            <div class="summary-card">
                <h3>Total Gastos</h3>
                <div class="value">{summary['total_gastos']}</div>
            </div>
            <div class="summary-card">
                <h3>Monto Total</h3>
                <div class="value">${summary['monto_total']:,.2f}</div>
            </div>
            <div class="summary-card">
                <h3>Promedio</h3>
                <div class="value">${summary['monto_promedio']:,.2f}</div>
            </div>
            <div class="summary-card">
                <h3>Monto Máximo</h3>
                <div class="value">${summary['monto_maximo']:,.2f}</div>
            </div>
        </div>
        """
    
    def _generate_gastos_table_html(self, gastos: List[Gasto], 
                                   options: ExportOptions) -> str:
        """Genera HTML de la tabla de gastos."""
        html = """
        <h2>Detalle de Gastos</h2>
        <table>
            <thead>
                <tr>
                    <th>Fecha</th>
                    <th>Hora</th>
                    <th>Monto</th>
                    <th>Categoría</th>
                    <th>Descripción</th>
                </tr>
            </thead>
            <tbody>
"""
        
        total = 0
        for gasto in gastos:
            total += float(gasto.monto)
            html += f"""
                <tr>
                    <td>{gasto.fecha.strftime(options.date_format)}</td>
                    <td>{gasto.fecha.strftime("%H:%M:%S")}</td>
                    <td class="amount">${float(gasto.monto):,.2f}</td>
                    <td>{gasto.categoria}</td>
                    <td>{gasto.descripcion or ""}</td>
                </tr>
"""
        
        # Fila de total
        html += f"""
                <tr class="total-row">
                    <td colspan="2"><strong>TOTAL</strong></td>
                    <td class="amount"><strong>${total:,.2f}</strong></td>
                    <td colspan="2"></td>
                </tr>
"""
        
        html += """
            </tbody>
        </table>
"""
        
        return html
    
    def _get_period_text(self, gastos: List[Gasto], options: ExportOptions) -> str:
        """Obtiene texto del período de gastos."""
        if not gastos:
            return "N/A"
        
        if options.filter_start_date and options.filter_end_date:
            return f"{options.filter_start_date.strftime(options.date_format)} - {options.filter_end_date.strftime(options.date_format)}"
        
        fechas = [g.fecha for g in gastos]
        fecha_min = min(fechas).strftime(options.date_format)
        fecha_max = max(fechas).strftime(options.date_format)
        
        return f"{fecha_min} - {fecha_max}"


class MultiFormatExporter:
    """Exportador multi-formato principal."""
    
    def __init__(self):
        self.config = get_config()
        self.logger = logger
        
        # Inicializar exportadores
        self.exporters = {
            ExportFormat.EXCEL: ExcelExporter(),
            ExportFormat.CSV: CSVExporter(),
            ExportFormat.JSON: JSONExporter(),
            ExportFormat.PDF: PDFExporter(),
            ExportFormat.HTML: HTMLExporter()
        }
    
    def export(self, gastos: List[Gasto], format: ExportFormat, 
               output_path: str, options: ExportOptions = None) -> ExportResult:
        """
        Exporta gastos al formato especificado.
        
        Args:
            gastos: Lista de gastos a exportar
            format: Formato de exportación
            output_path: Ruta del archivo de salida
            options: Opciones de exportación
            
        Returns:
            Resultado de la exportación
        """
        try:
            if format not in self.exporters:
                return ExportResult(
                    success=False,
                    error_message=f"Formato no soportado: {format.value}"
                )
            
            exporter = self.exporters[format]
            result = exporter.export(gastos, output_path, options)
            
            self.logger.info(f"Exportación {format.value} completada: {result.success}")
            return result
            
        except Exception as e:
            self.logger.error(f"Error en exportación {format.value}: {e}")
            return ExportResult(
                success=False,
                error_message=str(e)
            )
    
    def get_supported_formats(self) -> List[ExportFormat]:
        """
        Obtiene formatos soportados según dependencias instaladas.
        
        Returns:
            Lista de formatos soportados
        """
        supported = [ExportFormat.CSV, ExportFormat.JSON, ExportFormat.HTML]
        
        if HAS_OPENPYXL:
            supported.append(ExportFormat.EXCEL)
        
        if HAS_REPORTLAB:
            supported.append(ExportFormat.PDF)
        
        return supported


# Instancia global del exportador
_multi_exporter: Optional[MultiFormatExporter] = None


def get_exporter() -> MultiFormatExporter:
    """Obtiene instancia global del exportador."""
    global _multi_exporter
    if _multi_exporter is None:
        _multi_exporter = MultiFormatExporter()
    return _multi_exporter


def export_gastos(gastos: List[Gasto], format: ExportFormat, 
                 output_path: str, options: ExportOptions = None) -> ExportResult:
    """
    Función de conveniencia para exportar gastos.
    
    Args:
        gastos: Lista de gastos
        format: Formato de exportación
        output_path: Ruta de salida
        options: Opciones de exportación
        
    Returns:
        Resultado de la exportación
    """
    exporter = get_exporter()
    return exporter.export(gastos, format, output_path, options)