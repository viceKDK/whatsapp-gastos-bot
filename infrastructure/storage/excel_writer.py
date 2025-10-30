"""
Implementación Storage para Excel

Guarda gastos en archivos Excel usando openpyxl.
"""

import os
from typing import List
from datetime import date, datetime
from pathlib import Path
from decimal import Decimal

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl es requerido para ExcelStorage. Instalar con: pip install openpyxl")

from domain.models.gasto import Gasto
from shared.logger import get_logger


logger = get_logger(__name__)


class ExcelStorage:
    """Implementación de storage usando archivos Excel."""
    
    # Configuración de columnas
    COLUMNAS = {
        'A': {'nombre': 'Fecha', 'ancho': 12},
        'B': {'nombre': 'Hora', 'ancho': 10},
        'C': {'nombre': 'Monto', 'ancho': 12},
        'D': {'nombre': 'Categoría', 'ancho': 15},
        'E': {'nombre': 'Descripción', 'ancho': 30}
    }
    
    def __init__(self, archivo_path: str):
        """
        Inicializa el storage de Excel.
        
        Args:
            archivo_path: Ruta al archivo Excel
        """
        self.archivo_path = Path(archivo_path)
        self.logger = logger
        
        # Asegurar que el directorio existe
        self.archivo_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Crear archivo si no existe
        if not self.archivo_path.exists():
            self._crear_archivo_inicial()
    
    def _crear_archivo_inicial(self) -> None:
        """Crea el archivo Excel inicial con headers."""
        try:
            self.logger.info(f"Creando archivo Excel inicial: {self.archivo_path}")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Gastos"
            
            # Configurar headers
            for col, config in self.COLUMNAS.items():
                ws[f"{col}1"] = config['nombre']
                ws[f"{col}1"].font = Font(bold=True)
                ws[f"{col}1"].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                ws[f"{col}1"].alignment = Alignment(horizontal="center")
                
                # Configurar ancho de columna
                ws.column_dimensions[col].width = config['ancho']
            
            # Guardar archivo
            wb.save(self.archivo_path)
            self.logger.info("Archivo Excel creado exitosamente")
            
        except Exception as e:
            self.logger.error(f"Error creando archivo Excel: {str(e)}")
            raise
    
    def guardar_gasto(self, gasto: Gasto) -> bool:
        """
        Guarda un gasto en el archivo Excel.
        
        Args:
            gasto: Gasto a guardar
            
        Returns:
            True si se guardó exitosamente, False si no
        """
        try:
            self.logger.debug(f"Guardando gasto en Excel: {gasto}")
            
            # Cargar workbook
            wb = load_workbook(self.archivo_path)
            ws = wb.active
            
            # Encontrar próxima fila vacía
            fila = ws.max_row + 1
            
            # Escribir datos
            ws[f"A{fila}"] = gasto.fecha.strftime("%Y-%m-%d")
            ws[f"B{fila}"] = gasto.fecha.strftime("%H:%M:%S")
            ws[f"C{fila}"] = float(gasto.monto)
            ws[f"D{fila}"] = gasto.categoria
            ws[f"E{fila}"] = gasto.descripcion or ""
            
            # Aplicar formato a la fila
            self._aplicar_formato_fila(ws, fila)
            
            # Guardar archivo
            wb.save(self.archivo_path)
            
            # Asignar ID basado en número de fila (simplificado)
            gasto.id = fila - 1
            
            self.logger.info(f"Gasto guardado en Excel, fila {fila}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error guardando gasto en Excel: {str(e)}")
            return False
    
    def obtener_gastos(self, fecha_desde: date, fecha_hasta: date) -> List[Gasto]:
        """
        Obtiene gastos en un rango de fechas.
        
        Args:
            fecha_desde: Fecha inicio
            fecha_hasta: Fecha fin
            
        Returns:
            Lista de gastos en el rango
        """
        try:
            self.logger.debug(f"Obteniendo gastos desde {fecha_desde} hasta {fecha_hasta}")
            
            if not self.archivo_path.exists():
                self.logger.warning("Archivo Excel no existe")
                return []
            
            wb = load_workbook(self.archivo_path, data_only=True)
            ws = wb.active
            
            gastos = []
            
            # Leer filas (comenzar desde fila 2 para saltar headers)
            for fila in range(2, ws.max_row + 1):
                try:
                    fecha_str = ws[f"A{fila}"].value
                    hora_str = ws[f"B{fila}"].value
                    monto = ws[f"C{fila}"].value
                    categoria = ws[f"D{fila}"].value
                    descripcion = ws[f"E{fila}"].value
                    
                    # Validar que tengamos datos mínimos
                    if not all([fecha_str, monto, categoria]):
                        continue
                    
                    # Parsear fecha y hora
                    if isinstance(fecha_str, str):
                        fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                    else:
                        fecha_obj = fecha_str.date() if hasattr(fecha_str, 'date') else fecha_str
                    
                    # Filtrar por rango de fechas
                    if not (fecha_desde <= fecha_obj <= fecha_hasta):
                        continue
                    
                    # Construir datetime completo
                    if isinstance(hora_str, str):
                        hora_obj = datetime.strptime(hora_str, "%H:%M:%S").time()
                        fecha_completa = datetime.combine(fecha_obj, hora_obj)
                    else:
                        fecha_completa = datetime.combine(fecha_obj, hora_str) if hora_str else datetime.combine(fecha_obj, datetime.min.time())
                    
                    # Crear gasto
                    gasto = Gasto(
                        monto=Decimal(str(monto)),
                        categoria=str(categoria),
                        fecha=fecha_completa,
                        descripcion=str(descripcion) if descripcion else None
                    )
                    gasto.id = fila - 1
                    
                    gastos.append(gasto)
                    
                except Exception as e:
                    self.logger.warning(f"Error procesando fila {fila}: {str(e)}")
                    continue
            
            self.logger.info(f"Obtenidos {len(gastos)} gastos del Excel")
            return gastos
            
        except Exception as e:
            self.logger.error(f"Error obteniendo gastos del Excel: {str(e)}")
            return []

    def obtener_gastos_por_fecha(self, fecha_desde: datetime, fecha_hasta: datetime) -> List[Gasto]:
        """
        Obtiene gastos en un rango de fechas usando datetime.

        Args:
            fecha_desde: Fecha/hora de inicio
            fecha_hasta: Fecha/hora de fin

        Returns:
            Lista de gastos en el rango
        """
        try:
            return self.obtener_gastos(fecha_desde.date(), fecha_hasta.date())
        except Exception:
            return []

    def obtener_todos_gastos(self) -> List[Gasto]:
        """
        Obtiene todos los gastos registrados.

        Returns:
            Lista de todos los gastos
        """
        try:
            fecha_inicio = date(2000, 1, 1)
            fecha_fin = date(2100, 12, 31)
            return self.obtener_gastos(fecha_inicio, fecha_fin)
        except Exception:
            return []
    
    def obtener_gastos_por_categoria(self, categoria: str) -> List[Gasto]:
        """
        Obtiene todos los gastos de una categoría.
        
        Args:
            categoria: Nombre de la categoría
            
        Returns:
            Lista de gastos de la categoría
        """
        try:
            # Obtener todos los gastos (rango amplio)
            fecha_inicio = date(2000, 1, 1)
            fecha_fin = date(2100, 12, 31)
            
            todos_gastos = self.obtener_gastos(fecha_inicio, fecha_fin)
            
            # Filtrar por categoría
            gastos_categoria = [
                gasto for gasto in todos_gastos
                if gasto.categoria.lower() == categoria.lower()
            ]
            
            self.logger.info(f"Encontrados {len(gastos_categoria)} gastos en categoría '{categoria}'")
            return gastos_categoria
            
        except Exception as e:
            self.logger.error(f"Error obteniendo gastos por categoría: {str(e)}")
            return []
    
    def _aplicar_formato_fila(self, worksheet, fila: int) -> None:
        """
        Aplica formato a una fila de datos.
        
        Args:
            worksheet: Worksheet de openpyxl
            fila: Número de fila
        """
        try:
            # Formato para monto (moneda)
            worksheet[f"C{fila}"].number_format = '$#,##0.00'
            
            # Alineación
            for col in ['A', 'B', 'D']:
                worksheet[f"{col}{fila}"].alignment = Alignment(horizontal="center")
            
            worksheet[f"C{fila}"].alignment = Alignment(horizontal="right")
            worksheet[f"E{fila}"].alignment = Alignment(horizontal="left")
            
        except Exception as e:
            self.logger.warning(f"Error aplicando formato a fila {fila}: {str(e)}")
    
    def backup_archivo(self) -> str:
        """
        Crea un backup del archivo Excel.
        
        Returns:
            Ruta del archivo backup creado
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = self.archivo_path.with_name(f"{self.archivo_path.stem}_backup_{timestamp}{self.archivo_path.suffix}")
            
            if self.archivo_path.exists():
                import shutil
                shutil.copy2(self.archivo_path, backup_path)
                self.logger.info(f"Backup creado: {backup_path}")
                return str(backup_path)
            
            return ""
            
        except Exception as e:
            self.logger.error(f"Error creando backup: {str(e)}")
            return ""
    
    def obtener_estadisticas(self) -> dict:
        """
        Obtiene estadísticas del archivo Excel.
        
        Returns:
            Diccionario con estadísticas
        """
        try:
            fecha_inicio = date(2000, 1, 1)
            fecha_fin = date(2100, 12, 31)
            todos_gastos = self.obtener_gastos(fecha_inicio, fecha_fin)
            
            if not todos_gastos:
                return {'total_gastos': 0, 'monto_total': 0, 'categorias': []}
            
            from collections import Counter
            
            categorias = Counter(gasto.categoria for gasto in todos_gastos)
            monto_total = sum(gasto.monto for gasto in todos_gastos)
            
            return {
                'total_gastos': len(todos_gastos),
                'monto_total': float(monto_total),
                'categorias': dict(categorias),
                'fecha_primer_gasto': min(gasto.fecha for gasto in todos_gastos).isoformat(),
                'fecha_ultimo_gasto': max(gasto.fecha for gasto in todos_gastos).isoformat()
            }
            
        except Exception as e:
            self.logger.error(f"Error obteniendo estadísticas: {str(e)}")
            return {'total_gastos': 0, 'monto_total': 0, 'categorias': []}
