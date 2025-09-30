#!/usr/bin/env python3
"""
Script para eliminar gastos con timestamps futuros del Excel
"""

import sys
import os
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

def clean_future_gastos():
    """Elimina gastos con timestamps futuros del Excel."""
    
    excel_path = "data/gastos.xlsx"
    
    print("=== LIMPIANDO GASTOS FUTUROS DEL EXCEL ===")
    
    if not os.path.exists(excel_path):
        print(f"Excel no existe: {excel_path}")
        return
    
    try:
        # Cargar Excel
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Obtener timestamp actual
        now = datetime.now()
        today_str = now.date().isoformat()
        current_time = now.time()
        
        print(f"Fecha actual: {today_str}")
        print(f"Hora actual: {current_time.strftime('%H:%M:%S')}")
        
        # Buscar y eliminar filas con timestamps futuros
        rows_to_delete = []
        
        # Empezar desde fila 2 (fila 1 son headers)
        for row_num in range(2, ws.max_row + 1):
            fecha_cell = ws[f'A{row_num}']
            hora_cell = ws[f'B{row_num}']
            monto_cell = ws[f'C{row_num}']
            categoria_cell = ws[f'D{row_num}']
            
            # Saltar filas vacías
            if not fecha_cell.value or not hora_cell.value:
                continue
            
            # Verificar si es de hoy
            if isinstance(fecha_cell.value, datetime):
                fecha_gasto = fecha_cell.value.date()
            else:
                # Si es string, parsear
                try:
                    fecha_gasto = datetime.fromisoformat(str(fecha_cell.value)).date()
                except:
                    continue
            
            # Solo procesar gastos de hoy
            if fecha_gasto != now.date():
                continue
            
            # Verificar hora
            if isinstance(hora_cell.value, datetime):
                hora_gasto = hora_cell.value.time()
            else:
                try:
                    hora_gasto = datetime.strptime(str(hora_cell.value), '%H:%M:%S').time()
                except:
                    continue
            
            # Crear datetime completo del gasto
            gasto_datetime = datetime.combine(fecha_gasto, hora_gasto)
            
            # Si es futuro, marcar para eliminación
            if gasto_datetime > now:
                monto = monto_cell.value if monto_cell.value else "N/A"
                categoria = categoria_cell.value if categoria_cell.value else "N/A"
                
                print(f"FUTURO DETECTADO - Fila {row_num}: {gasto_datetime} - ${monto} - {categoria}")
                rows_to_delete.append(row_num)
        
        # Eliminar filas (de abajo hacia arriba para no alterar índices)
        if rows_to_delete:
            print(f"\nEliminando {len(rows_to_delete)} filas con timestamps futuros...")
            
            for row_num in sorted(rows_to_delete, reverse=True):
                print(f"  Eliminando fila {row_num}")
                ws.delete_rows(row_num)
            
            # Guardar Excel
            wb.save(excel_path)
            print(f"✅ Excel actualizado: {len(rows_to_delete)} gastos futuros eliminados")
        else:
            print("✅ No hay gastos futuros para eliminar")
        
        wb.close()
        
    except Exception as e:
        print(f"ERROR: {e}")

if __name__ == "__main__":
    clean_future_gastos()